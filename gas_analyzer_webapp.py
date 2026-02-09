"""
Gas Turbine Fuel Analyzer - Clean Web App
Professional web application with Google Sheets integration
"""

import streamlit as st
import pandas as pd
import json
import math
from datetime import datetime
from dataclasses import dataclass
import io

# Google Sheets (optional - will work without it)
try:
    import gspread
    from google.oauth2.service_account import Credentials
    SHEETS_OK = True
except:
    SHEETS_OK = False

# Excel export (optional)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except:
    EXCEL_OK = False

# PDF export (optional)
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    PDF_OK = True
except:
    PDF_OK = False

# Page setup
st.set_page_config(
    page_title="Gas Analyzer Pro",
    page_icon="🔥",
    layout="wide"
)

# CSS Styling
st.markdown("""
<style>
    .main {background-color: #f8f9fa;}
    .stButton button {
        background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%);
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.5rem 2rem;
        border: none;
    }
    h1 {color: #1e3a8a;}
    h2 {color: #3b82f6;}
</style>
""", unsafe_allow_html=True)

# Component data
@dataclass
class Component:
    name: str
    formula: str
    mw: float
    lhv: float
    hhv: float

COMPONENTS = {
    'Methane': Component('Methane', 'CH4', 16.043, 50.01, 55.50),
    'Ethane': Component('Ethane', 'C2H6', 30.070, 47.49, 51.88),
    'Propane': Component('Propane', 'C3H8', 44.097, 46.35, 50.36),
    'n-Butane': Component('n-Butane', 'C4H10', 58.123, 45.75, 49.50),
    'i-Butane': Component('i-Butane', 'C4H10', 58.123, 45.61, 49.36),
    'n-Pentane': Component('n-Pentane', 'C5H12', 72.150, 45.36, 49.01),
    'i-Pentane': Component('i-Pentane', 'C5H12', 72.150, 45.24, 48.89),
    'n-Hexane': Component('n-Hexane', 'C6H14', 86.177, 45.10, 48.68),
    'Heptane': Component('Heptane', 'C7H16', 100.204, 44.93, 48.45),
    'Hydrogen': Component('Hydrogen', 'H2', 2.016, 120.00, 141.80),
    'Carbon Monoxide': Component('Carbon Monoxide', 'CO', 28.010, 10.10, 10.10),
    'Carbon Dioxide': Component('Carbon Dioxide', 'CO2', 44.010, 0.00, 0.00),
    'Nitrogen': Component('Nitrogen', 'N2', 28.014, 0.00, 0.00),
    'Hydrogen Sulfide': Component('Hydrogen Sulfide', 'H2S', 34.081, 15.20, 16.53),
}

PRESETS = {
    'Pipeline Natural Gas': {
        'Methane': 95.0, 'Ethane': 2.5, 'Propane': 0.5, 
        'n-Butane': 0.2, 'Carbon Dioxide': 1.0, 'Nitrogen': 0.8
    },
    'Rich Natural Gas': {
        'Methane': 85.0, 'Ethane': 8.0, 'Propane': 4.0,
        'n-Butane': 1.5, 'Carbon Dioxide': 0.5, 'Nitrogen': 1.0
    },
    'Lean Natural Gas': {
        'Methane': 98.0, 'Ethane': 0.5, 
        'Carbon Dioxide': 1.0, 'Nitrogen': 0.5
    },
}

DEFAULT_LIMITS = {
    'wobbe_lower': {'min': 47, 'max': 51, 'name': 'Wobbe Index (L)'},
    'lhv_vol': {'min': 32, 'max': 40, 'name': 'LHV (volume)'},
    'sg': {'min': 0.55, 'max': 0.75, 'name': 'Specific Gravity'},
    'mn': {'min': 80, 'max': 999, 'name': 'Methane Number'},
    'h2': {'min': 0, 'max': 5, 'name': 'H2 Content'},
    'co2_n2': {'min': 0, 'max': 10, 'name': 'Inerts'},
    'h2s': {'min': 0, 'max': 5, 'name': 'H2S Content'},
}

# Session state
if 'composition' not in st.session_state:
    st.session_state.composition = {name: 0.0 for name in COMPONENTS.keys()}
if 'results' not in st.session_state:
    st.session_state.results = {}
if 'use_si' not in st.session_state:
    st.session_state.use_si = True
if 'limits' not in st.session_state:
    st.session_state.limits = DEFAULT_LIMITS.copy()

# Callback functions for preset loading
def load_preset_callback():
    """Load selected preset into composition"""
    selected = st.session_state.get('preset_selector', 'Custom')
    if selected != 'Custom':
        # Clear all first
        for name in COMPONENTS.keys():
            st.session_state.composition[name] = 0.0
        # Load preset
        for name, value in PRESETS[selected].items():
            st.session_state.composition[name] = value

def clear_all_callback():
    """Clear all composition values"""
    for name in COMPONENTS.keys():
        st.session_state.composition[name] = 0.0

def calculate_properties(comp_percent):
    """Calculate all gas properties from composition"""
    comp = {k: v/100 for k, v in comp_percent.items() if v > 0}
    
    if not comp:
        return None
    
    total = sum(comp.values())
    comp = {k: v/total for k, v in comp.items()}
    
    # Basic properties
    mw = sum(comp[n] * COMPONENTS[n].mw for n in comp)
    sg = mw / 28.97
    dens_si = mw / 22.414
    dens_us = mw / 379.49
    
    # Heating values
    lhv_m_si = sum((comp[n] * COMPONENTS[n].mw / mw) * COMPONENTS[n].lhv for n in comp)
    hhv_m_si = sum((comp[n] * COMPONENTS[n].mw / mw) * COMPONENTS[n].hhv for n in comp)
    lhv_v_si = lhv_m_si * dens_si
    hhv_v_si = hhv_m_si * dens_si
    
    # Wobbe Index
    wi_l_si = lhv_v_si / math.sqrt(sg)
    wi_h_si = hhv_v_si / math.sqrt(sg)
    
    # Advanced properties
    h2 = comp.get('Hydrogen', 0) * 100
    co2_n2 = (comp.get('Carbon Dioxide', 0) + comp.get('Nitrogen', 0)) * 100
    h2s = comp.get('Hydrogen Sulfide', 0) * 1e6
    
    mn = (137.78 * comp.get('Methane', 0) - 
          40 * comp.get('Ethane', 0) - 
          79.52 * comp.get('Propane', 0) + 
          1.5 * co2_n2/100)
    
    o2 = (comp.get('Methane', 0) * 2 + 
          comp.get('Ethane', 0) * 3.5 + 
          comp.get('Propane', 0) * 5 + 
          comp.get('Hydrogen', 0) * 0.5)
    afr = (o2 / 0.2095 * 28.97) / mw
    
    aft_c = 1900 + (lhv_v_si / 40) * 100 - (co2_n2 * 15)
    aft_f = aft_c * 1.8 + 32
    
    lel = 0
    fsi = comp.get('Methane', 0) * 1.0 + comp.get('Ethane', 0) * 0.9
    
    return {
        'composition': comp,
        'mw': mw, 'sg': sg, 'dens_si': dens_si, 'dens_us': dens_us,
        'lhv_m_si': lhv_m_si, 'lhv_v_si': lhv_v_si,
        'lhv_m_us': lhv_m_si * 429.923, 'lhv_v_us': lhv_m_si * 429.923 * dens_us,
        'hhv_m_si': hhv_m_si, 'hhv_v_si': hhv_v_si,
        'hhv_m_us': hhv_m_si * 429.923, 'hhv_v_us': hhv_m_si * 429.923 * dens_us,
        'wi_l_si': wi_l_si, 'wi_h_si': wi_h_si,
        'wi_l_us': wi_l_si * 26.839, 'wi_h_us': wi_h_si * 26.839,
        'h2': h2, 'co2_n2': co2_n2, 'h2s': h2s, 'mn': mn, 
        'afr': afr, 'aft_c': aft_c, 'aft_f': aft_f, 'lel': lel, 'fsi': fsi
    }

def generate_pdf_report(project, source, analyst, comp_input, results, limits, use_si):
    """Generate branded PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    elements.append(Paragraph("GAS TURBINE FUEL QUALITY ANALYSIS", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Project info
    info_data = [
        ['Project:', project, 'Date:', datetime.now().strftime('%Y-%m-%d')],
        ['Source:', source, 'Analyst:', analyst],
        ['Units:', 'SI (Metric)' if use_si else 'US (Imperial)', '', '']
    ]
    info_table = Table(info_data, colWidths=[1*inch, 2*inch, 1*inch, 1.5*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#1e3a8a')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Overall status
    r = results
    si = use_si
    lims = limits
    
    checks = [
        ('wobbe_lower', 'wi_l_si'), ('lhv_vol', 'lhv_v_si'), ('sg', 'sg'),
        ('mn', 'mn'), ('h2', 'h2'), ('co2_n2', 'co2_n2')
    ]
    
    statuses = [check_status(k, r[rk], lims) for k, rk in checks if k in lims]
    overall = "NOT SUITABLE" if 'FAIL' in statuses else "SUITABLE FOR TURBINE USE"
    overall_color = colors.red if 'FAIL' in statuses else colors.green
    
    overall_para = Paragraph(f"<b>OVERALL STATUS: {overall}</b>", 
                            ParagraphStyle('overall', fontSize=16, textColor=overall_color, alignment=TA_CENTER))
    elements.append(overall_para)
    elements.append(Spacer(1, 0.2*inch))
    
    # Gas Composition
    elements.append(Paragraph("GAS COMPOSITION", heading_style))
    comp_data = [['Component', 'Formula', 'Mol%']]
    sorted_comp = sorted(r['composition'].items(), key=lambda x: x[1], reverse=True)
    for name, frac in sorted_comp:
        comp_data.append([name, COMPONENTS[name].formula, f"{frac*100:.2f}%"])
    
    comp_table = Table(comp_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    comp_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(comp_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Key Properties
    elements.append(Paragraph("CALCULATED PROPERTIES", heading_style))
    props_data = [['Property', 'Value', 'Unit', 'Status']]
    
    for lim_key, res_key in checks:
        if lim_key in lims:
            lim = lims[lim_key]
            
            if lim_key == 'wobbe_lower':
                val = r['wi_l_si'] if si else r['wi_l_us']
                unit = 'MJ/m3' if si else 'Btu/scf'
            elif lim_key == 'lhv_vol':
                val = r['lhv_v_si'] if si else r['lhv_v_us']
                unit = 'MJ/m3' if si else 'Btu/scf'
            else:
                val = r[res_key]
                unit = 'mol%' if lim_key in ['h2', 'co2_n2'] else ('-' if lim_key in ['sg', 'mn'] else 'ppmv')
            
            status = check_status(lim_key, r[res_key] if lim_key not in ['wobbe_lower', 'lhv_vol'] else r[res_key], lims)
            props_data.append([lim['name'], f"{val:.2f}", unit, status])
    
    props_table = Table(props_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
    props_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(props_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Footer/Disclaimer
    disclaimer = Paragraph(
        "<i>This report is for informational purposes only. Consult turbine OEM specifications for final approval.</i>",
        ParagraphStyle('disclaimer', fontSize=8, textColor=colors.grey)
    )
    elements.append(Spacer(1, 0.3*inch))
    elements.append(disclaimer)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

# SIDEBAR
with st.sidebar:
    st.title("Settings")
    
    st.session_state.use_si = st.radio(
        "Units",
        [True, False],
        format_func=lambda x: "SI (Metric)" if x else "US (Imperial)",
        index=0 if st.session_state.use_si else 1
    )
    
    st.markdown("---")
    st.markdown("### About")
    st.markdown("""
    **Gas Analyzer Pro v2.0**
    
    Professional fuel analysis for gas turbines.
    
    - Complete property calculations
    - SI and US units
    - Excel export
    - Custom limits
    """)

# MAIN APP
st.title("Gas Turbine Fuel Quality Analyzer")
st.markdown("*Professional Web Application*")

tabs = st.tabs(["Input", "Results", "Summary", "Settings"])

# TAB 1: INPUT
with tabs[0]:
    col1, col2, col3 = st.columns(3)
    with col1:
        project = st.text_input("Project Name")
    with col2:
        source = st.text_input("Gas Source")
    with col3:
        analyst = st.text_input("Analyst")
    
    st.markdown("### Gas Composition")
    st.markdown("**Enter mol% for each component:**")
    
    col1, col2 = st.columns(2)
    comp_input = {}
    
    components_list = list(COMPONENTS.items())
    mid = len(components_list) // 2
    
    with col1:
        for name, comp in components_list[:mid]:
            # Get value from session state composition
            comp_input[name] = st.number_input(
                f"{name} ({comp.formula})",
                min_value=0.0,
                max_value=100.0,
                value=float(st.session_state.composition.get(name, 0.0)),
                step=0.1,
                format="%.2f",
                key=f"inp_{name}"
            )
            # Update session state as user types
            st.session_state.composition[name] = comp_input[name]
    
    with col2:
        for name, comp in components_list[mid:]:
            # Get value from session state composition
            comp_input[name] = st.number_input(
                f"{name} ({comp.formula})",
                min_value=0.0,
                max_value=100.0,
                value=float(st.session_state.composition.get(name, 0.0)),
                step=0.1,
                format="%.2f",
                key=f"inp_{name}"
            )
            # Update session state as user types
            st.session_state.composition[name] = comp_input[name]
    
    total = sum(comp_input.values())
    if abs(total - 100) < 0.1:
        st.success(f"Total: {total:.2f}%")
    elif total > 0:
        st.warning(f"Total: {total:.2f}% (Should be 100%)")
    
    if st.button("CALCULATE PROPERTIES", type="primary", use_container_width=True):
        results = calculate_properties(comp_input)
        if results:
            st.session_state.results = results
            st.session_state.composition = comp_input
            st.success("Calculation complete! Check Results tab.")
        else:
            st.error("Invalid composition")

# TAB 2: RESULTS
with tabs[1]:
    if not st.session_state.results:
        st.info("Enter composition and calculate first")
    else:
        r = st.session_state.results
        si = st.session_state.use_si
        
        st.subheader("Gas Composition")
        comp_df = pd.DataFrame([
            {
                'Component': name,
                'Formula': COMPONENTS[name].formula,
                'Mol%': f"{frac*100:.2f}%"
            }
            for name, frac in sorted(r['composition'].items(), 
                                    key=lambda x: x[1], reverse=True)
        ])
        st.dataframe(comp_df, use_container_width=True, hide_index=True)
        
        st.subheader("Calculated Properties")
        
        props = [
            ['Molecular Weight', f"{r['mw']:.3f}", 'g/mol' if si else 'lb/lbmol'],
            ['Specific Gravity', f"{r['sg']:.4f}", '-'],
            ['Density', f"{r['dens_si' if si else 'dens_us']:.4f}", 
             'kg/m3' if si else 'lb/ft3'],
            ['LHV (mass)', f"{r['lhv_m_si' if si else 'lhv_m_us']:.2f}", 
             'MJ/kg' if si else 'Btu/lb'],
            ['LHV (volume)', f"{r['lhv_v_si' if si else 'lhv_v_us']:.2f}", 
             'MJ/m3' if si else 'Btu/scf'],
            ['HHV (mass)', f"{r['hhv_m_si' if si else 'hhv_m_us']:.2f}", 
             'MJ/kg' if si else 'Btu/lb'],
            ['HHV (volume)', f"{r['hhv_v_si' if si else 'hhv_v_us']:.2f}", 
             'MJ/m3' if si else 'Btu/scf'],
            ['Wobbe Index (L)', f"{r['wi_l_si' if si else 'wi_l_us']:.2f}", 
             'MJ/m3' if si else 'Btu/scf'],
            ['Wobbe Index (H)', f"{r['wi_h_si' if si else 'wi_h_us']:.2f}", 
             'MJ/m3' if si else 'Btu/scf'],
            ['H2 Content', f"{r['h2']:.2f}", 'mol%'],
            ['CO2+N2', f"{r['co2_n2']:.2f}", 'mol%'],
            ['H2S', f"{r['h2s']:.1f}", 'ppmv'],
            ['Methane Number', f"{r['mn']:.1f}", '-'],
            ['Air/Fuel Ratio', f"{r['afr']:.2f}", 'kg/kg' if si else 'lb/lb'],
            ['Flame Temp', f"{r['aft_c' if si else 'aft_f']:.0f}", 
             'C' if si else 'F'],
        ]
        
        props_df = pd.DataFrame(props, columns=['Property', 'Value', 'Unit'])
        st.dataframe(props_df, use_container_width=True, hide_index=True)
        
        if EXCEL_OK:
            if st.button("Download Excel Report"):
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Gas Analysis Report"
                ws['A1'].font = Font(bold=True, size=14)
                
                ws['A3'] = f"Project: {project}"
                ws['A4'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
                
                ws['A6'] = "Property"
                ws['B6'] = "Value"
                ws['C6'] = "Unit"
                
                for i, prop in enumerate(props, start=7):
                    ws[f'A{i}'] = prop[0]
                    ws[f'B{i}'] = prop[1]
                    ws[f'C{i}'] = prop[2]
                
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    "Download Excel",
                    output,
                    f"gas_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# TAB 3: SUMMARY
with tabs[2]:
    if not st.session_state.results:
        st.info("Calculate properties first")
    else:
        r = st.session_state.results
        si = st.session_state.use_si
        lims = st.session_state.limits
        
        checks = [
            ('wobbe_lower', 'wi_l_si'),
            ('lhv_vol', 'lhv_v_si'),
            ('sg', 'sg'),
            ('mn', 'mn'),
            ('h2', 'h2'),
            ('co2_n2', 'co2_n2')
        ]
        
        statuses = []
        for lim_key, res_key in checks:
            if lim_key in lims:
                status = check_status(lim_key, r[res_key], lims)
                statuses.append(status)
        
        if 'FAIL' in statuses:
            st.error("NOT SUITABLE FOR TURBINE USE")
        else:
            st.success("SUITABLE FOR TURBINE USE")
        
        st.markdown("---")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Wobbe Index", 
                     f"{r['wi_l_si' if si else 'wi_l_us']:.2f}")
            st.metric("LHV", 
                     f"{r['lhv_v_si' if si else 'lhv_v_us']:.2f}")
        with col2:
            st.metric("Specific Gravity", f"{r['sg']:.4f}")
            st.metric("Methane Number", f"{r['mn']:.0f}")
        with col3:
            st.metric("H2 Content", f"{r['h2']:.2f}%")
            st.metric("Inerts", f"{r['co2_n2']:.2f}%")
        
        st.subheader("Detailed Assessment")
        
        status_data = []
        for lim_key, res_key in checks:
            if lim_key in lims:
                lim = lims[lim_key]
                
                # Get value with proper unit conversion
                if lim_key == 'wobbe_lower':
                    val = r['wi_l_si'] if si else r['wi_l_us']
                    check_val = r['wi_l_si']
                    unit = 'MJ/m3' if si else 'Btu/scf'
                    range_min = lim['min'] if si else lim['min'] * 26.839
                    range_max = lim['max'] if si else lim['max'] * 26.839
                elif lim_key == 'lhv_vol':
                    val = r['lhv_v_si'] if si else r['lhv_v_us']
                    check_val = r['lhv_v_si']
                    unit = 'MJ/m3' if si else 'Btu/scf'
                    range_min = lim['min'] if si else lim['min'] * 26.839
                    range_max = lim['max'] if si else lim['max'] * 26.839
                else:
                    val = r[res_key]
                    check_val = val
                    unit = 'mol%' if lim_key in ['h2', 'co2_n2'] else ('-' if lim_key in ['sg', 'mn'] else 'ppmv')
                    range_min = lim['min']
                    range_max = lim['max']
                
                status = check_status(lim_key, check_val, lims)
                
                status_data.append({
                    'Property': lim['name'],
                    'Value': f"{val:.2f} {unit}",
                    'Range': f"{range_min:.0f}-{range_max:.0f} {unit}",
                    'Status': status
                })
        
        st.dataframe(
            pd.DataFrame(status_data),
            use_container_width=True,
            hide_index=True
        )
        
        # Report Generation
        st.markdown("---")
        st.subheader("📄 Generate Report")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if PDF_OK:
                if st.button("📄 Download PDF Report", use_container_width=True, type="primary"):
                    pdf_buffer = generate_pdf_report(
                        project, source, analyst, comp_input, r, lims, si
                    )
                    st.download_button(
                        "💾 Download PDF",
                        pdf_buffer,
                        f"gas_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                        "application/pdf",
                        use_container_width=True
                    )
            else:
                st.info("Install reportlab for PDF: pip install reportlab")
        
        with col2:
            if EXCEL_OK:
                if st.button("📊 Download Excel Report", use_container_width=True):
                    # Enhanced Excel with branding
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Gas Analysis"
                    
                    # Header
                    ws['A1'] = "GAS TURBINE FUEL QUALITY ANALYSIS"
                    ws['A1'].font = Font(bold=True, size=16, color="1E3A8A")
                    ws.merge_cells('A1:D1')
                    ws['A1'].alignment = Alignment(horizontal='center')
                    
                    # Project info
                    ws['A3'] = "Project:"
                    ws['B3'] = project
                    ws['C3'] = "Date:"
                    ws['D3'] = datetime.now().strftime('%Y-%m-%d')
                    ws['A4'] = "Source:"
                    ws['B4'] = source
                    ws['C4'] = "Analyst:"
                    ws['D4'] = analyst
                    
                    # Composition
                    ws['A6'] = "GAS COMPOSITION"
                    ws['A6'].font = Font(bold=True, size=12)
                    ws['A7'] = "Component"
                    ws['B7'] = "Mol%"
                    
                    for cell in ['A7', 'B7']:
                        ws[cell].font = Font(bold=True)
                        ws[cell].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                    
                    row = 8
                    for name, frac in sorted(r['composition'].items(), key=lambda x: x[1], reverse=True):
                        ws[f'A{row}'] = name
                        ws[f'B{row}'] = f"{frac*100:.2f}%"
                        row += 1
                    
                    wb.save(output)
                    output.seek(0)
                    
                    st.download_button(
                        "💾 Download Excel",
                        output,
                        f"gas_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

# TAB 4: SETTINGS
with tabs[3]:
    st.subheader("Custom Acceptance Limits")
    
    for key, lim in st.session_state.limits.items():
        st.markdown(f"**{lim['name']}**")
        col1, col2 = st.columns(2)
        with col1:
            new_min = st.number_input(
                f"Min {lim['name']}", 
                value=float(lim['min']), 
                key=f"min_{key}"
            )
        with col2:
            new_max = st.number_input(
                f"Max {lim['name']}", 
                value=float(lim['max']), 
                key=f"max_{key}"
            )
        
        st.session_state.limits[key]['min'] = new_min
        st.session_state.limits[key]['max'] = new_max
    
    if st.button("Reset to Defaults"):
        st.session_state.limits = DEFAULT_LIMITS.copy()
        st.rerun()

st.markdown("---")
st.markdown("*Gas Turbine Fuel Analyzer Pro - Web Edition*")
